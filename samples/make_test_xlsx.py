"""
テスト用 Excel ファイルを生成するスクリプト。

生成するファイル:
  1. simple.xlsx         - 単純な表（結合なし）
  2. merged_cells.xlsx   - 結合セル（横・縦・矩形）を含む表
  3. multi_block.xlsx    - 空行で区切られた複数ブロック（タイトル行 + データ表）
  4. large_rows.xlsx     - max_rows 超過チェック用（ヘッダー + 25 データ行）
"""

import openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path

OUT = Path(__file__).parent / "xlsx"
OUT.mkdir(exist_ok=True)


# ------------------------------------------------------------------ #
# 1. simple.xlsx  単純なヘッダー + データ行
# ------------------------------------------------------------------ #
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "売上一覧"

ws.append(["日付", "商品名", "数量", "単価", "合計"])
ws.append(["2024-01-05", "リンゴ", 10, 150, 1500])
ws.append(["2024-01-06", "バナナ", 20, 80, 1600])
ws.append(["2024-01-07", "ミカン", 15, 120, 1800])

wb.save(OUT / "simple.xlsx")
print("✓ simple.xlsx")


# ------------------------------------------------------------------ #
# 2. merged_cells.xlsx  横結合・縦結合・矩形結合
# ------------------------------------------------------------------ #
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "結合テスト"

#  A1:D1 横結合タイトル
ws["A1"] = "第1四半期 売上報告書"
ws.merge_cells("A1:D1")

# ヘッダー行
ws.append(["地域", "1月", "2月", "3月"])

# A3:A5 縦結合（地域ラベル）
ws["A3"] = "東日本"
ws["B3"] = 100
ws["C3"] = 120
ws["D3"] = 110
ws["A4"] = ""        # 縦結合継続
ws["B4"] = 90
ws["C4"] = 80
ws["D4"] = 95
ws["A5"] = ""        # 縦結合継続
ws["B5"] = 70
ws["C5"] = 60
ws["D5"] = 75
ws.merge_cells("A3:A5")

# B6:C7 矩形結合
ws["A6"] = "備考"
ws["B6"] = "特記事項なし"
ws.merge_cells("B6:C7")

wb.save(OUT / "merged_cells.xlsx")
print("✓ merged_cells.xlsx")


# ------------------------------------------------------------------ #
# 3. multi_block.xlsx  空行区切りの複数ブロック
#    → elements に Paragraph + Table + Table が出力されることを確認
# ------------------------------------------------------------------ #
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "申請書"

# ブロック1: タイトル（1行1セル → Paragraph）
ws["A1"] = "年次有給休暇申請書"

# 空行
# row 2 は空のまま

# ブロック2: メタ情報（複数セル1行 → Table）
ws["A3"] = "申請者"
ws["B3"] = "山田 太郎"
ws["A4"] = "所属"
ws["B4"] = "総務部"
ws["A5"] = "申請日"
ws["B5"] = "2024-03-01"

# 空行
# row 6 は空のまま

# ブロック3: 詳細テーブル（複数行 → Table）
ws.append(["", "", "", "", "", "", ""])  # row 7 空
ws["A7"] = ""  # 空行として機能させる（既に空）
ws["A8"] = "休暇開始日"
ws["B8"] = "休暇終了日"
ws["C8"] = "日数"
ws["D8"] = "理由"
ws["A9"] = "2024-03-15"
ws["B9"] = "2024-03-17"
ws["C9"] = 3
ws["D9"] = "私用"
ws["A10"] = "2024-04-01"
ws["B10"] = "2024-04-01"
ws["C10"] = 1
ws["D10"] = "私用"

wb.save(OUT / "multi_block.xlsx")
print("✓ multi_block.xlsx")


# ------------------------------------------------------------------ #
# 4. large_rows.xlsx  max_rows 超過 → 子 Section 分割確認
# ------------------------------------------------------------------ #
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "大量データ"

ws.append(["ID", "名前", "スコア"])
for i in range(1, 26):  # 25 データ行（デフォルト max_rows=20 を超える）
    ws.append([i, f"ユーザー{i:02d}", i * 3])

wb.save(OUT / "large_rows.xlsx")
print("✓ large_rows.xlsx")

print(f"\n生成先: {OUT}")
